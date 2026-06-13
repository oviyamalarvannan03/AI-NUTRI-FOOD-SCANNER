import os
import openpyxl

def parse_summary(ws):
    summary_dict = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row and row[0] is not None:
            summary_dict[str(row[0]).strip()] = row[1]
    return summary_dict

def parse_details(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) for h in rows[0]]
    details = []
    for r in rows[1:]:
        if r and r[0] is not None:
            details.append(dict(zip(headers, r)))
    return headers, details

def main():
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    tests_dir = os.path.dirname(os.path.abspath(__file__))
    e2e_path = os.path.join(tests_dir, "reports", "Appium_Android_E2E_Test_Report.xlsx")
    sec_path = os.path.join(tests_dir, "reports", "Vulnerability_Security_Report.xlsx")
    
    # Fallback to root just in case
    if not os.path.exists(e2e_path):
        e2e_path = os.path.join(os.path.dirname(tests_dir), "Appium_Android_E2E_Test_Report.xlsx")
    if not os.path.exists(sec_path):
        sec_path = os.path.join(os.path.dirname(tests_dir), "Vulnerability_Security_Report.xlsx")

    e2e_wb = openpyxl.load_workbook(e2e_path, data_only=True)
    sec_wb = openpyxl.load_workbook(sec_path, data_only=True)
    
    e2e_summary = parse_summary(e2e_wb['Summary'])
    sec_summary = parse_summary(sec_wb['Summary'])
    
    e2e_headers, e2e_details = parse_details(e2e_wb['Test Results'])
    sec_headers, sec_details = parse_details(sec_wb['Security Audit Results'])
    
    markdown_output = []
    markdown_output.append("# 🧪 NutriAI Automated Test Verification Dashboard\n")
    markdown_output.append("This dashboard displays the test results verified from the completed test execution reports.\n")
    
    # E2E Test Suite Summary
    markdown_output.append("## 📱 E2E Appium Android Test Suite Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Application Under Test** | {e2e_summary.get('Application Under Test', 'N/A')} |")
    markdown_output.append(f"| **Simulation Presets** | {e2e_summary.get('Simulation Presets', 'N/A')} |")
    markdown_output.append(f"| **Target Platform** | {e2e_summary.get('Target Platform', 'N/A')} |")
    markdown_output.append(f"| **Total Test Cases** | {e2e_summary.get('Total Tests', 'N/A')} |")
    markdown_output.append(f"| **Passed** | ✅ {e2e_summary.get('Passed', 'N/A')} |")
    markdown_output.append(f"| **Failed** | ❌ {e2e_summary.get('Failed', 'N/A')} |")
    markdown_output.append(f"| **Pass Rate** | **{e2e_summary.get('Pass Rate', 'N/A')}** |")
    markdown_output.append(f"| **Test Date** | {e2e_summary.get('Test Date', 'N/A')} |")
    markdown_output.append("\n")
    
    # Security Vulnerability Summary
    markdown_output.append("## 🛡️ Backend & Client-Side Security Verification Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Target Application** | {sec_summary.get('Target Application', 'N/A')} |")
    markdown_output.append(f"| **Audited Host URL** | {sec_summary.get('Audited Host URL', 'N/A')} |")
    markdown_output.append(f"| **Audit Type** | {sec_summary.get('Audit Type', 'N/A')} |")
    markdown_output.append(f"| **Total Test Cases Checked** | {sec_summary.get('Total Test Cases Checked', 'N/A')} |")
    markdown_output.append(f"| **Passed** | ✅ {sec_summary.get('Passed', 'N/A')} |")
    markdown_output.append(f"| **Failed (Vulnerabilities)** | ❌ {sec_summary.get('Failed (Vulnerabilities Found)', 'N/A')} |")
    markdown_output.append(f"| **Remediation Status** | **{sec_summary.get('Remediation Status', 'N/A')}** |")
    markdown_output.append(f"| **Audit Date** | {sec_summary.get('Audit Date', 'N/A')} |")
    markdown_output.append("\n")
    
    # E2E Details Expandable Section
    num_e2e = len(e2e_details)
    markdown_output.append("### 📋 E2E Test Cases Detail Breakdowns")
    markdown_output.append(f"<details><summary>Click to view all E2E Test Cases ({num_e2e} tests)</summary>\n")
    markdown_output.append("| Test Case ID | Description | Status | Notes / Error | Timestamp |")
    markdown_output.append("|---|---|---|---|---|")
    for r in e2e_details:
        status_raw = r.get('Status', '')
        status_emoji = "✅ PASSED" if str(status_raw).strip().lower() in ("pass", "passed") else "❌ FAILED"
        markdown_output.append(f"| `{r.get('Test Case ID')}` | {r.get('Description')} | {status_emoji} | {r.get('Notes / Error') or ''} | {r.get('Timestamp')} |")
    markdown_output.append("\n</details>\n")
    
    # Security Details Expandable Section
    num_sec = len(sec_details)
    markdown_output.append("### 🔐 Security Test Cases Detail Breakdowns")
    markdown_output.append(f"<details><summary>Click to view all Security Test Cases ({num_sec} tests)</summary>\n")
    markdown_output.append("| Test Case ID | Vulnerability Type | File Path | Severity | Status |")
    markdown_output.append("|---|---|---|---|---|")
    for r in sec_details:
        status_raw = r.get('Status', '')
        status_emoji = "✅ PASSED" if str(status_raw).strip().lower() in ("pass", "passed") else "❌ FAILED"
        markdown_output.append(f"| `{r.get('Test Case ID')}` | {r.get('Vulnerability Type')} | `{r.get('File Path')}` | {r.get('Severity')} | {status_emoji} |")
    markdown_output.append("\n</details>\n")
    
    markdown_output.append("## 📦 Downloadable Test Report Artifacts")
    markdown_output.append("The full Excel spreadsheets (`.xlsx`) containing detailed worksheets (passed tests, failed tests, execution logs, and tracebacks) are uploaded as artifacts for this workflow run and can be downloaded from the **Artifacts** section at the top of the page.")
    
    full_markdown = "\n".join(markdown_output)
    
    # Write to GITHUB_STEP_SUMMARY
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "w", encoding="utf-8") as f:
            f.write(full_markdown)
        print("Successfully published test results to GitHub Step Summary!")
    else:
        print(full_markdown)

if __name__ == "__main__":
    main()
