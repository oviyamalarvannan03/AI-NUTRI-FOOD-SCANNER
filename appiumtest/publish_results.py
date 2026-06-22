import os
import openpyxl

def parse_summary_sheet(sheet):
    summary = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and row[0] is not None:
            summary[str(row[0]).strip()] = str(row[1]).strip()
    return summary

def parse_details_sheet(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() for h in rows[0]]
    details = []
    for r in rows[1:]:
        if r and r[0] is not None:
            details.append(dict(zip(headers, r)))
    return headers, details

def main():
    # Configure UTF-8 stdout if possible to prevent Windows encoding crashes when printing emojis
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    # Paths are relative to the root of the repository
    repo_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    e2e_path = os.environ.get("REPORT_FILE", os.path.join(repo_dir, "Selenium_E2E_Test_Report.xlsx"))
    if not os.path.isabs(e2e_path):
        e2e_path = os.path.join(repo_dir, e2e_path)
        
    sec_path = os.path.join(repo_dir, "Vulnerability_Security_Report.xlsx")
    
    # Parse Selenium E2E Report
    wb_e2e = openpyxl.load_workbook(e2e_path, data_only=True)
    e2e_summary = parse_summary_sheet(wb_e2e['Summary'])
    e2e_headers, e2e_details = parse_details_sheet(wb_e2e['Test Results'])
    
    # Calculate stats for E2E
    e2e_total = len(e2e_details)
    e2e_passed = sum(1 for r in e2e_details if str(r.get('Status')).strip().lower() == 'pass')
    e2e_failed = e2e_total - e2e_passed
    e2e_pass_rate = round((e2e_passed / e2e_total) * 100, 2) if e2e_total > 0 else 0.0

    # Parse Security Audit Report
    wb_sec = openpyxl.load_workbook(sec_path, data_only=True)
    sec_summary = parse_summary_sheet(wb_sec['Summary'])
    sec_headers, sec_details = parse_details_sheet(wb_sec['Security Audit Results'])
    
    # Calculate stats for Security Audit
    sec_total = len(sec_details)
    sec_passed = sum(1 for r in sec_details if str(r.get('Status')).strip().lower() == 'pass')
    sec_failed = sec_total - sec_passed
    sec_pass_rate = round((sec_passed / sec_total) * 100, 2) if sec_total > 0 else 0.0

    import datetime
    current_timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    markdown_output = []
    markdown_output.append("# 🧪 NutriAI Automated Test & Security Verification Dashboard\n")
    markdown_output.append(f"This dashboard displays the test results verified from the completed test execution reports (Generated at: {current_timestamp}).\n")
    
    # E2E Test Suite Summary
    markdown_output.append("## 🌿 Web E2E Test Suite Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Application Under Test** | {e2e_summary.get('Application Under Test')} |")
    markdown_output.append(f"| **Test URL** | {e2e_summary.get('Test URL')} |")
    markdown_output.append(f"| **Target Platform** | {e2e_summary.get('Target Platform')} |")
    markdown_output.append(f"| **Test Date** | {e2e_summary.get('Test Date')} |")
    markdown_output.append(f"| **Total Test Cases** | {e2e_total} |")
    markdown_output.append(f"| **Passed** | ✅ {e2e_passed} |")
    markdown_output.append(f"| **Failed** | ❌ {e2e_failed} |")
    markdown_output.append(f"| **Pass Rate** | **{e2e_pass_rate}%** |")
    markdown_output.append("\n")
    
    # Security Vulnerability Summary
    markdown_output.append("## 🛡️ Backend Security Audit Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Target Application** | {sec_summary.get('Target Application')} |")
    markdown_output.append(f"| **Audited Host URL** | {sec_summary.get('Audited Host URL')} |")
    markdown_output.append(f"| **Audit Type** | {sec_summary.get('Audit Type')} |")
    markdown_output.append(f"| **Audit Date** | {sec_summary.get('Audit Date')} |")
    markdown_output.append(f"| **Total Security Checks** | {sec_total} |")
    markdown_output.append(f"| **Passed** | ✅ {sec_passed} |")
    markdown_output.append(f"| **Failed** | ❌ {sec_failed} |")
    markdown_output.append(f"| **Pass Rate** | **{sec_pass_rate}%** |")
    markdown_output.append("\n")
    
    # E2E Details Expandable Section
    markdown_output.append("### 📋 E2E Test Cases Detail Breakdowns")
    markdown_output.append(f"<details><summary>Click to view all E2E Test Cases ({e2e_total} tests)</summary>\n")
    markdown_output.append("| Test Case ID | Description | Status | Notes / Error | Timestamp |")
    markdown_output.append("|---|---|---|---|---|")
    for r in e2e_details:
        status_emoji = "✅ Pass" if str(r.get("Status")).strip().lower() == "pass" else "❌ Fail"
        markdown_output.append(f"| {r.get('Test Case ID')} | {r.get('Description')} | {status_emoji} | {r.get('Notes / Error') or ''} | {r.get('Timestamp') or ''} |")
    markdown_output.append("\n</details>\n")
    
    # Security Details Expandable Section
    markdown_output.append("### 🔐 Security Test Cases Detail Breakdowns")
    markdown_output.append(f"<details><summary>Click to view all Security Test Cases ({sec_total} tests)</summary>\n")
    markdown_output.append("| Test Case ID | Vulnerability Type | File Path | Severity | Explanation | Remediation | Status |")
    markdown_output.append("|---|---|---|---|---|---|---|")
    for r in sec_details:
        status_emoji = "✅ Pass" if str(r.get("Status")).strip().lower() == "pass" else "❌ Fail"
        markdown_output.append(f"| {r.get('Test Case ID')} | {r.get('Vulnerability Type')} | `{r.get('File Path')}` | {r.get('Severity')} | {r.get('Explanation')} | {r.get('Remediation')} | {status_emoji} |")
    markdown_output.append("\n</details>\n")
    
    markdown_output.append("## 📦 Downloadable Test Report Artifacts")
    markdown_output.append("The full Excel spreadsheets (`.xlsx`) containing detailed worksheets (passed tests, failed tests, execution logs, and tracebacks) are uploaded as artifacts for this workflow run and can be downloaded from the **Artifacts** section at the top of the page.")
    
    full_markdown = "\n".join(markdown_output)
    
    # Write to GITHUB_STEP_SUMMARY
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "w", encoding="utf-8") as f:
            f.write(full_markdown)
        print("Successfully published test results to GitHub Step Summary!")
    else:
        print(full_markdown)

if __name__ == "__main__":
    main()
